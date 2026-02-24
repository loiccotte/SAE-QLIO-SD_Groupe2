"""
Export des KPIs en fichiers Excel (.xlsx) et PDF.

Ce module fournit deux routes d'export, accessibles uniquement aux
utilisateurs ayant le role ``responsable`` ou superieur :

- ``/export/excel`` : classeur Excel avec 5 onglets (Performance,
  Qualite, Delai, Energie, Stock), mis en forme avec openpyxl.
- ``/export/pdf`` : rapport PDF genere via weasyprint (avec fallback
  HTML si weasyprint n'est pas installe).

Le nom du fichier telecharge suit le format :
``telefan_kpis_YYYYMMDD_HHMM.{xlsx|pdf|html}``
"""

import io
import logging
from datetime import datetime

from flask import Blueprint, request, send_file

from . import services
from .auth import login_required, role_required

bp = Blueprint('export', __name__)

logger = logging.getLogger(__name__)


# ============================================================================
# Fonctions utilitaires
# ============================================================================

def _collect_kpis() -> dict:
    """Collecte l'ensemble des 11 KPIs pour l'export.

    Returns:
        Dictionnaire {nom_kpi: dict_resultat} pour tous les KPIs.
    """
    return {
        'oee': services.calculate_oee(),
        'utilization': services.calculate_utilization(),
        'throughput': services.calculate_throughput(),
        'cycle_time': services.calculate_cycle_time(),
        'non_conformity': services.calculate_non_conformity(),
        'detection_time': services.calculate_detection_time(),
        'lead_time': services.calculate_lead_time(),
        'buffer_wait': services.calculate_buffer_wait_time(),
        'energy': services.calculate_energy_summary(),
        'buffer_occupancy': services.calculate_buffer_occupancy(),
        'stock_variation': services.calculate_stock_variation(),
    }


def _get_filters() -> dict[str, str]:
    """Recupere les filtres temporels depuis les parametres de requete.

    Returns:
        Dictionnaire {year, month, day, hour} (chaines vides si absent).
    """
    return {
        'year': request.args.get('year', ''),
        'month': request.args.get('month', ''),
        'day': request.args.get('day', ''),
        'hour': request.args.get('hour', ''),
    }


def _filters_label(filters: dict[str, str]) -> str:
    """Genere un label descriptif des filtres temporels appliques.

    Args:
        filters: Dictionnaire retourne par ``_get_filters()``.

    Returns:
        Chaine descriptive, ex: 'Annee: 2025 | Mois: 03'
        ou 'Toutes les donnees' si aucun filtre.
    """
    parts = []
    if filters['year']:
        parts.append(f"Annee: {filters['year']}")
    if filters['month']:
        parts.append(f"Mois: {filters['month']}")
    if filters['day']:
        parts.append(f"Jour: {filters['day']}")
    if filters['hour']:
        parts.append(f"Heure: {filters['hour']}")
    return ' | '.join(parts) if parts else 'Toutes les donnees'


def _generate_filename(extension: str) -> str:
    """Genere un nom de fichier horodate pour l'export.

    Args:
        extension: Extension du fichier (sans point), ex: 'xlsx', 'pdf'.

    Returns:
        Nom de fichier, ex: 'telefan_kpis_20250327_1430.xlsx'.
    """
    return f"telefan_kpis_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"


# ============================================================================
# Export Excel
# ============================================================================

@bp.route('/export/excel')
@login_required
@role_required('responsable')
def export_excel():
    """Genere et telecharge un classeur Excel avec 5 onglets KPI.

    Onglets :
    1. Performance — OEE, cadence, temps de cycle, utilisation par machine
    2. Qualite — Non-conformite, temps de detection, detail par machine
    3. Delai — Lead time, distribution par ordre
    4. Energie — Consommation electrique et air comprime, timeline
    5. Stock — Occupation buffers, variation de stock
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    kpis = _collect_kpis()
    filters = _get_filters()
    label = _filters_label(filters)
    wb = Workbook()

    # --- Styles communs ---
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Couleurs d'en-tete par categorie KPI
    HEADER_FILLS = {
        'perf':    PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'qualite': PatternFill(start_color='38B6FF', end_color='38B6FF', fill_type='solid'),
        'delai':   PatternFill(start_color='F2C0FF', end_color='F2C0FF', fill_type='solid'),
        'energie': PatternFill(start_color='09B200', end_color='09B200', fill_type='solid'),
        'stock':   PatternFill(start_color='737373', end_color='737373', fill_type='solid'),
    }

    def _write_header(ws, row, headers, fill):
        """Ecrit une ligne d'en-tete stylisee dans la feuille."""
        for col, header_text in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header_text)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def _write_data_row(ws, row, values):
        """Ecrit une ligne de donnees avec bordures."""
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value).border = thin_border

    # --- Onglet 1 : Performance ---
    ws = wb.active
    ws.title = 'Performance'
    ws.cell(row=1, column=1, value=f"T'ELEFAN MES 4.0 - Export {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Filtre: {label}").font = Font(italic=True, size=10)

    ws.cell(row=4, column=1, value='OEE Global').font = Font(bold=True)
    ws.cell(row=4, column=2, value=f"{kpis['oee'].get('value', 0)}%")
    ws.cell(row=5, column=1, value='Disponibilite')
    ws.cell(row=5, column=2, value=f"{kpis['oee'].get('availability', 0)}%")
    ws.cell(row=6, column=1, value='Performance')
    ws.cell(row=6, column=2, value=f"{kpis['oee'].get('performance', 0)}%")
    ws.cell(row=7, column=1, value='Qualite')
    ws.cell(row=7, column=2, value=f"{kpis['oee'].get('quality', 0)}%")

    ws.cell(row=9, column=1, value='Cadence reelle').font = Font(bold=True)
    ws.cell(row=9, column=2, value=f"{kpis['throughput'].get('value', 0)} pieces/h")
    ws.cell(row=10, column=1, value='Temps cycle moyen').font = Font(bold=True)
    ws.cell(row=10, column=2, value=f"{kpis['cycle_time'].get('value', 0)} s")

    row = 12
    _write_header(ws, row, ['Machine', 'Utilisation (%)'], HEADER_FILLS['perf'])
    for m in kpis['utilization'].get('by_machine', []):
        row += 1
        _write_data_row(ws, row, [m['name'], m['value']])

    # --- Onglet 2 : Qualite ---
    ws2 = wb.create_sheet('Qualite')
    ws2.cell(row=1, column=1, value=f"Filtre: {label}").font = Font(italic=True, size=10)
    ws2.cell(row=3, column=1, value='Taux non-conformite').font = Font(bold=True)
    ws2.cell(row=3, column=2, value=f"{kpis['non_conformity'].get('value', 0)}%")
    ws2.cell(row=4, column=1, value='Taux ordres')
    ws2.cell(row=4, column=2, value=f"{kpis['non_conformity'].get('rate_orders', 0)}%")
    ws2.cell(row=5, column=1, value='Taux detection pieces')
    ws2.cell(row=5, column=2, value=f"{kpis['non_conformity'].get('rate_parts', 0)}%")
    ws2.cell(row=7, column=1, value='Temps detection moyen').font = Font(bold=True)
    ws2.cell(row=7, column=2, value=f"{kpis['detection_time'].get('value', 0)} s")

    row = 9
    _write_header(ws2, row, ['Machine', 'Total', 'Erreurs', 'Taux (%)'], HEADER_FILLS['qualite'])
    for m in kpis['non_conformity'].get('by_machine', []):
        row += 1
        _write_data_row(ws2, row, [m['name'], m['total'], m['errors'], m['rate']])

    # --- Onglet 3 : Delai ---
    ws3 = wb.create_sheet('Delai')
    ws3.cell(row=1, column=1, value=f"Filtre: {label}").font = Font(italic=True, size=10)
    ws3.cell(row=3, column=1, value='Lead Time moyen').font = Font(bold=True)
    ws3.cell(row=3, column=2, value=f"{kpis['lead_time'].get('value', 0)} h")
    ws3.cell(row=4, column=1, value='Temps attente buffer moyen').font = Font(bold=True)
    ws3.cell(row=4, column=2, value=f"{kpis['buffer_wait'].get('value', 0)} s")

    row = 6
    _write_header(ws3, row, ['Ordre', 'Lead Time (h)', 'Debut'], HEADER_FILLS['delai'])
    for d in kpis['lead_time'].get('distribution', []):
        row += 1
        _write_data_row(ws3, row, [d['order'], d['hours'], d['start']])

    # --- Onglet 4 : Energie ---
    ws4 = wb.create_sheet('Energie')
    ws4.cell(row=1, column=1, value=f"Filtre: {label}").font = Font(italic=True, size=10)
    ws4.cell(row=3, column=1, value='Energie par unite').font = Font(bold=True)
    ws4.cell(row=3, column=2, value=f"{kpis['energy'].get('value', 0)} {kpis['energy'].get('unit', 'Wh/u')}")
    ws4.cell(row=4, column=1, value='Air comprime par unite').font = Font(bold=True)
    ws4.cell(row=4, column=2, value=f"{kpis['energy'].get('air_value', 0)} {kpis['energy'].get('air_unit', 'L/u')}")
    ws4.cell(row=5, column=1, value='Note')
    ws4.cell(row=5, column=2, value=kpis['energy'].get('note', ''))

    row = 7
    _write_header(ws4, row, ['Periode', 'Consommation (Wh)'], HEADER_FILLS['energie'])
    for t in kpis['energy'].get('timeline', []):
        row += 1
        _write_data_row(ws4, row, [t['period'], t['kwh']])

    # --- Onglet 5 : Stock ---
    ws5 = wb.create_sheet('Stock')
    ws5.cell(row=1, column=1, value=f"Filtre: {label}").font = Font(italic=True, size=10)
    ws5.cell(row=3, column=1, value='Taux occupation global').font = Font(bold=True)
    ws5.cell(row=3, column=2, value=f"{kpis['buffer_occupancy'].get('value', 0)}%")

    row = 5
    _write_header(ws5, row, ['Buffer', 'Capacite', 'Occupe', 'Taux (%)'], HEADER_FILLS['stock'])
    for b in kpis['buffer_occupancy'].get('by_buffer', []):
        row += 1
        _write_data_row(ws5, row, [b['name'], b['capacity'], b['occupied'], b['rate']])

    row += 2
    _write_header(ws5, row, ['Buffer', 'Variation (%)'], HEADER_FILLS['stock'])
    for v in kpis['stock_variation'].get('variations', []):
        row += 1
        _write_data_row(ws5, row, [v['buffer'], v['variation_pct']])

    # Ajustement automatique de la largeur des colonnes
    for ws_sheet in [ws, ws2, ws3, ws4, ws5]:
        for col in ws_sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_sheet.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Serialisation et envoi du fichier
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=_generate_filename('xlsx'),
    )


# ============================================================================
# Export PDF
# ============================================================================

@bp.route('/export/pdf')
@login_required
@role_required('responsable')
def export_pdf():
    """Genere et telecharge un rapport PDF des KPIs.

    Le PDF est construit a partir d'un template HTML inline, puis
    converti via weasyprint. Si weasyprint n'est pas disponible,
    le fichier HTML est renvoye directement en tant que fallback.
    """
    kpis = _collect_kpis()
    filters = _get_filters()
    label = _filters_label(filters)
    now = datetime.now().strftime('%d/%m/%Y %H:%M')

    html_content = _build_pdf_html(kpis, label, now)

    # Tentative de conversion weasyprint, sinon fallback HTML
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        output = io.BytesIO(pdf_bytes)
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=_generate_filename('pdf'),
        )
    except ImportError:
        logger.warning("weasyprint non installe, fallback vers export HTML.")
        output = io.BytesIO(html_content.encode('utf-8'))
        return send_file(
            output,
            mimetype='text/html',
            as_attachment=True,
            download_name=_generate_filename('html'),
        )


def _build_pdf_html(kpis: dict, label: str, now: str) -> str:
    """Construit le HTML du rapport PDF.

    Args:
        kpis: Dictionnaire complet des KPIs (retour de ``_collect_kpis``).
        label: Label descriptif des filtres temporels.
        now: Date/heure courante formatee.

    Returns:
        Chaine HTML complete du rapport.
    """
    parts = [
        f"""<html><head><meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; color: #18181b; }}
            h1 {{ font-size: 22px; margin-bottom: 4px; }}
            h2 {{ font-size: 16px; color: #555; border-bottom: 2px solid #ddd; padding-bottom: 4px; margin-top: 24px; }}
            .meta {{ font-size: 11px; color: #888; margin-bottom: 20px; }}
            table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 6px 10px; font-size: 12px; }}
            th {{ background-color: #f4f4f5; font-weight: bold; text-align: left; }}
            .kpi-value {{ font-size: 28px; font-weight: bold; }}
            .section {{ margin-bottom: 20px; }}
        </style></head><body>
        <h1>T'ELEFAN MES 4.0 - Rapport KPIs</h1>
        <p class="meta">Genere le {now} | Filtre: {label}</p>"""
    ]

    # Section Performance
    parts.append('<h2 style="color: #ff0000;">Performance</h2>')
    parts.append('<div class="section">')
    parts.append(f'<p>OEE Global: <span class="kpi-value">{kpis["oee"].get("value", 0)}%</span></p>')
    parts.append(f'<p>Disponibilite: {kpis["oee"].get("availability", 0)}% | '
                 f'Performance: {kpis["oee"].get("performance", 0)}% | '
                 f'Qualite: {kpis["oee"].get("quality", 0)}%</p>')
    parts.append(f'<p>Cadence: {kpis["throughput"].get("value", 0)} pieces/h | '
                 f'Temps cycle: {kpis["cycle_time"].get("value", 0)} s</p>')
    if kpis['utilization'].get('by_machine'):
        parts.append('<table><tr><th>Machine</th><th>Utilisation (%)</th></tr>')
        for m in kpis['utilization']['by_machine']:
            parts.append(f'<tr><td>{m["name"]}</td><td>{m["value"]}</td></tr>')
        parts.append('</table>')
    parts.append('</div>')

    # Section Qualite
    parts.append('<h2 style="color: #38b6ff;">Qualite</h2>')
    parts.append('<div class="section">')
    parts.append(f'<p>Taux non-conformite: <span class="kpi-value">{kpis["non_conformity"].get("value", 0)}%</span></p>')
    parts.append(f'<p>Temps detection moyen: {kpis["detection_time"].get("value", 0)} s</p>')
    parts.append('</div>')

    # Section Delai
    parts.append('<h2 style="color: #d946a8;">Delai</h2>')
    parts.append('<div class="section">')
    parts.append(f'<p>Lead Time moyen: <span class="kpi-value">{kpis["lead_time"].get("value", 0)} h</span></p>')
    parts.append(f'<p>Temps attente buffer: {kpis["buffer_wait"].get("value", 0)} s '
                 f'({kpis["buffer_wait"].get("count", 0)} observations)</p>')
    parts.append('</div>')

    # Section Energie
    parts.append('<h2 style="color: #09b200;">Energie</h2>')
    parts.append('<div class="section">')
    parts.append(f'<p>Consommation: <span class="kpi-value">{kpis["energy"].get("value", 0)} {kpis["energy"].get("unit", "Wh/u")}</span></p>')
    parts.append(f'<p>Air comprime: {kpis["energy"].get("air_value", 0)} {kpis["energy"].get("air_unit", "L/u")}</p>')
    parts.append(f'<p><em>{kpis["energy"].get("note", "")}</em></p>')
    parts.append('</div>')

    # Section Stock
    parts.append('<h2 style="color: #737373;">Stock</h2>')
    parts.append('<div class="section">')
    parts.append(f'<p>Occupation globale: <span class="kpi-value">{kpis["buffer_occupancy"].get("value", 0)}%</span></p>')
    if kpis['buffer_occupancy'].get('by_buffer'):
        parts.append('<table><tr><th>Buffer</th><th>Capacite</th><th>Occupe</th><th>Taux (%)</th></tr>')
        for b in kpis['buffer_occupancy']['by_buffer']:
            style = ' style="color: red; font-weight: bold;"' if b['rate'] > 90 else ''
            parts.append(f'<tr><td>{b["name"]}</td><td>{b["capacity"]}</td>'
                         f'<td>{b["occupied"]}</td><td{style}>{b["rate"]}</td></tr>')
        parts.append('</table>')
    parts.append('</div>')

    parts.append('</body></html>')
    return '\n'.join(parts)
